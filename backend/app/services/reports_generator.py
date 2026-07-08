"""
Gerador de relatórios PDF (via Gotenberg) e XLSX (via openpyxl).
Tipos: financeiro | projetos | propostas
"""
import calendar as cal_mod
from datetime import date as date_cls
from pathlib import Path
from typing import Optional

import httpx

MONTHS_PT = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]


# ── Helpers ───────────────────────────────────────────────────────────────────

def _fmt(v: float) -> str:
    return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _row(i: int) -> str:
    return "#F8FAFC" if i % 2 == 0 else "#FFFFFF"


def _pct(a: float, b: float) -> str:
    return f"{a/b*100:.1f}%" if b > 0 else "—"


def _d(s: Optional[str]) -> str:
    if not s:
        return "—"
    try:
        y, m, d = str(s).split("T")[0].split("-")
        return f"{d}/{m}/{y}"
    except Exception:
        return str(s)


def _base_css() -> str:
    return """
      *{margin:0;padding:0;box-sizing:border-box;font-family:'Helvetica Neue',Arial,sans-serif}
      body{padding:36px;color:#1E293B;font-size:12px;line-height:1.5}
      .hdr{display:flex;justify-content:space-between;align-items:flex-start;
           border-bottom:3px solid #1D4ED8;padding-bottom:16px;margin-bottom:24px}
      .logo h1{font-size:20px;font-weight:800;color:#1D4ED8}
      .logo p{color:#64748B;font-size:11px;margin-top:2px}
      .badge{background:#1D4ED8;color:#fff;padding:5px 14px;border-radius:5px;
             font-size:10px;font-weight:700;letter-spacing:1px;align-self:flex-start}
      .sec{margin-bottom:20px}
      .sec-title{font-size:10px;font-weight:700;color:#1D4ED8;text-transform:uppercase;
                 letter-spacing:1px;margin-bottom:8px;padding-bottom:4px;border-bottom:1px solid #E2E8F0}
      .kpi-row{display:grid;gap:8px;margin-bottom:18px}
      .kpi{border-radius:8px;padding:12px 14px;color:#fff}
      .kpi label{font-size:9px;opacity:.8;display:block;margin-bottom:2px;
                 text-transform:uppercase;letter-spacing:.5px}
      .kpi span{font-size:17px;font-weight:800}
      .kpi.blue{background:linear-gradient(135deg,#1D4ED8,#2563EB)}
      .kpi.red{background:linear-gradient(135deg,#DC2626,#EF4444)}
      .kpi.green{background:linear-gradient(135deg,#059669,#10B981)}
      .kpi.amber{background:linear-gradient(135deg,#D97706,#F59E0B)}
      .kpi.violet{background:linear-gradient(135deg,#7C3AED,#8B5CF6)}
      .kpi.slate{background:linear-gradient(135deg,#475569,#64748B)}
      table{width:100%;border-collapse:collapse;font-size:11px}
      thead tr{background:#1D4ED8;color:#fff}
      thead th{padding:6px 10px;text-align:left;font-size:10px;font-weight:600;letter-spacing:.5px}
      tbody td{padding:5px 10px}
      .num{text-align:right;font-weight:600}
      .green{color:#059669}.red{color:#DC2626}.blue{color:#1D4ED8}
      .ftr{margin-top:28px;padding-top:12px;border-top:1px solid #E2E8F0;
           display:flex;justify-content:space-between;color:#94A3B8;font-size:10px}
    """


async def _gotenberg(gotenberg_url: str, html: str) -> bytes:
    async with httpx.AsyncClient(timeout=45) as client:
        r = await client.post(
            f"{gotenberg_url}/forms/chromium/convert/html",
            files={"files": ("index.html", html.encode("utf-8"), "text/html")},
        )
        r.raise_for_status()
    return r.content


def _save(storage: Path, name: str, content: bytes) -> str:
    storage.mkdir(parents=True, exist_ok=True)
    path = storage / name
    path.write_bytes(content)
    return str(path)


# ── FINANCEIRO ────────────────────────────────────────────────────────────────

async def pdf_financeiro(
    gotenberg_url: str, storage: Path,
    user_id: str, entries: list[dict],
    cat_map: dict[str, str],
    date_from: Optional[date_cls], date_to: Optional[date_cls],
) -> str:
    df  = date_from or date_cls.today().replace(day=1)
    dt  = date_to   or date_cls.today()
    period_label = f"{_d(df.isoformat())} a {_d(dt.isoformat())}"

    total_in  = sum(float(e["value"]) for e in entries if e.get("type") == "entrada")
    total_out = sum(float(e["value"]) for e in entries if e.get("type") == "saida")
    saldo     = total_in - total_out

    # Monthly breakdown
    monthly: dict[str, dict] = {}
    for e in entries:
        key = str(e.get("date",""))[:7]
        if not key:
            continue
        r = monthly.setdefault(key, {"ent": 0.0, "sai": 0.0})
        if e.get("type") == "entrada":
            r["ent"] += float(e["value"])
        else:
            r["sai"] += float(e["value"])

    monthly_rows = ""
    acc = 0.0
    for key in sorted(monthly):
        y, m = key.split("-")
        label = f"{MONTHS_PT[int(m)-1]}/{y}"
        ent = monthly[key]["ent"]
        sai = monthly[key]["sai"]
        sal = ent - sai
        acc += sal
        color = "#059669" if sal >= 0 else "#DC2626"
        monthly_rows += f"""<tr>
          <td style="padding:5px 10px">{label}</td>
          <td class="num" style="color:#059669;padding:5px 10px">{_fmt(ent)}</td>
          <td class="num" style="color:#DC2626;padding:5px 10px">{_fmt(sai)}</td>
          <td class="num" style="color:{color};padding:5px 10px">{_fmt(sal)}</td>
          <td class="num" style="padding:5px 10px">{_fmt(acc)}</td>
        </tr>"""

    # Expense categories
    exp: dict[str, float] = {}
    inc: dict[str, float] = {}
    for e in entries:
        cat = e.get("category") or "outro"
        if e.get("type") == "saida":
            exp[cat] = exp.get(cat, 0) + float(e["value"])
        else:
            inc[cat] = inc.get(cat, 0) + float(e["value"])

    def cat_rows(d: dict) -> str:
        rows = ""
        for i, (k, v) in enumerate(sorted(d.items(), key=lambda x: -x[1])):
            lbl = cat_map.get(k, k)
            pct = v / sum(d.values()) * 100 if d else 0
            rows += f'<tr style="background:{_row(i)}"><td style="padding:5px 10px">{lbl}</td>'
            rows += f'<td class="num" style="padding:5px 10px">{_fmt(v)}</td>'
            rows += f'<td class="num" style="padding:5px 10px">{pct:.1f}%</td></tr>'
        return rows

    # Entries table (limited)
    LIMIT = 80
    entries_rows = ""
    for i, e in enumerate(entries[:LIMIT]):
        cor = "#059669" if e.get("type") == "entrada" else "#DC2626"
        tipo = "Entrada" if e.get("type") == "entrada" else "Saída"
        cat = cat_map.get(e.get("category",""), e.get("category","—"))
        entries_rows += f"""<tr style="background:{_row(i)}">
          <td style="padding:5px 10px">{_d(e.get('date',''))}</td>
          <td style="padding:5px 10px;color:{cor};font-weight:600">{tipo}</td>
          <td style="padding:5px 10px">{cat}</td>
          <td style="padding:5px 10px">{(e.get('description') or e.get('supplier') or '—')[:50]}</td>
          <td class="num" style="padding:5px 10px">{_fmt(float(e.get('value',0)))}</td>
        </tr>"""
    extra_note = f'<p style="font-size:10px;color:#94A3B8;margin-top:6px">Exibindo {min(len(entries),LIMIT)} de {len(entries)} lançamentos.</p>' if len(entries) > LIMIT else ""

    html = f"""<!DOCTYPE html><html lang="pt-BR"><head><meta charset="UTF-8"/>
<style>
{_base_css()}
.kpi-row{{grid-template-columns:repeat(3,1fr)}}
</style></head><body>
<div class="hdr">
  <div class="logo"><h1>InfraReport</h1><p>Relatório Financeiro · {period_label}</p></div>
  <span class="badge">FINANCEIRO</span>
</div>

<div class="kpi-row">
  <div class="kpi green"><label>Total Entradas</label><span>{_fmt(total_in)}</span></div>
  <div class="kpi red"><label>Total Saídas</label><span>{_fmt(total_out)}</span></div>
  <div class="kpi {'green' if saldo >= 0 else 'red'}"><label>Saldo do Período</label><span>{_fmt(saldo)}</span></div>
</div>

<div class="sec"><div class="sec-title">Resumo Mensal</div>
<table><thead><tr><th>Mês</th><th style="text-align:right">Entradas</th>
<th style="text-align:right">Saídas</th><th style="text-align:right">Saldo</th>
<th style="text-align:right">Acumulado</th></tr></thead>
<tbody>{monthly_rows if monthly_rows else '<tr><td colspan="5" style="text-align:center;padding:12px;color:#94A3B8">Sem dados no período</td></tr>'}</tbody>
</table></div>

{'<div class="sec"><div class="sec-title">Saídas por Categoria</div><table><thead><tr><th>Categoria</th><th style="text-align:right">Total</th><th style="text-align:right">%</th></tr></thead><tbody>' + cat_rows(exp) + '</tbody></table></div>' if exp else ''}

{'<div class="sec"><div class="sec-title">Receitas por Tipo</div><table><thead><tr><th>Tipo</th><th style="text-align:right">Total</th><th style="text-align:right">%</th></tr></thead><tbody>' + cat_rows(inc) + '</tbody></table></div>' if inc else ''}

<div class="sec"><div class="sec-title">Lançamentos do Período</div>
<table><thead><tr><th>Data</th><th>Tipo</th><th>Categoria</th><th>Descrição / Fornecedor</th>
<th style="text-align:right">Valor</th></tr></thead>
<tbody>{entries_rows if entries_rows else '<tr><td colspan="5" style="text-align:center;padding:12px;color:#94A3B8">Nenhum lançamento no período</td></tr>'}</tbody>
</table>{extra_note}</div>

<div class="ftr"><span>InfraReport — Gestão de Infraestrutura Predial</span>
<span>Gerado em {date_cls.today().strftime("%d/%m/%Y")}</span></div>
</body></html>"""

    content = await _gotenberg(gotenberg_url, html)
    fname = f"financeiro_{df.strftime('%Y%m')}_{dt.strftime('%Y%m')}.pdf"
    return _save(storage, fname, content)


# ── PROJETOS ──────────────────────────────────────────────────────────────────

async def pdf_projetos(
    gotenberg_url: str, storage: Path,
    user_id: str, projects: list[dict],
    date_from: Optional[date_cls], date_to: Optional[date_cls],
) -> str:
    STATUS_PT = {"aguardando": "Aguardando", "em_andamento": "Em Andamento",
                 "concluido": "Concluído", "cancelado": "Cancelado"}

    total    = len(projects)
    active   = sum(1 for p in projects if p.get("status") == "em_andamento")
    done     = sum(1 for p in projects if p.get("status") == "concluido")
    canc     = sum(1 for p in projects if p.get("status") == "cancelado")
    tot_rev  = sum(float(p.get("revenue") or 0) for p in projects)
    tot_cost = sum(float(p.get("material_cost",0) or 0) + float(p.get("labor_cost",0) or 0) for p in projects if p.get("status") in ("em_andamento","concluido"))

    # Avg margin (completed only)
    done_margins = []
    for p in projects:
        if p.get("status") == "concluido":
            rev  = float(p.get("revenue") or 0)
            cost = float(p.get("material_cost") or 0) + float(p.get("labor_cost") or 0)
            if rev > 0:
                done_margins.append((rev - cost) / rev * 100)
    avg_margin = sum(done_margins) / len(done_margins) if done_margins else 0

    # Avg duration (days, completed only)
    durations = []
    for p in projects:
        if p.get("status") == "concluido" and p.get("start_date") and p.get("end_date"):
            try:
                sd = date_cls.fromisoformat(str(p["start_date"]).split("T")[0])
                ed = date_cls.fromisoformat(str(p["end_date"]).split("T")[0])
                durations.append((ed - sd).days)
            except Exception:
                pass
    avg_duration = sum(durations) / len(durations) if durations else 0

    STATUS_COLORS = {"aguardando": "#F59E0B", "em_andamento": "#3B82F6",
                     "concluido": "#10B981", "cancelado": "#EF4444"}
    proj_rows = ""
    for i, p in enumerate(projects):
        rev  = float(p.get("revenue") or 0)
        cost = float(p.get("material_cost",0) or 0) + float(p.get("labor_cost",0) or 0)
        mar  = (rev - cost) / rev * 100 if rev > 0 else 0
        sc   = STATUS_COLORS.get(p.get("status",""), "#64748B")
        proj_rows += f"""<tr style="background:{_row(i)}">
          <td style="padding:5px 10px;font-weight:600">{(p.get('name') or '—')[:35]}</td>
          <td style="padding:5px 10px">{(p.get('client') or '—')[:25]}</td>
          <td style="padding:5px 10px"><span style="background:{sc}22;color:{sc};padding:2px 8px;border-radius:99px;font-size:10px;font-weight:700">{STATUS_PT.get(p.get('status',''),'—')}</span></td>
          <td class="num" style="padding:5px 10px;color:#059669">{_fmt(rev)}</td>
          <td class="num" style="padding:5px 10px;color:#DC2626">{_fmt(cost)}</td>
          <td class="num" style="padding:5px 10px">{mar:.1f}%</td>
          <td style="padding:5px 10px">{_d(p.get('start_date'))}</td>
          <td style="padding:5px 10px">{_d(p.get('end_date') or p.get('expected_end_date'))}</td>
        </tr>"""

    html = f"""<!DOCTYPE html><html lang="pt-BR"><head><meta charset="UTF-8"/>
<style>
{_base_css()}
.kpi-row{{grid-template-columns:repeat(4,1fr)}}
</style></head><body>
<div class="hdr">
  <div class="logo"><h1>InfraReport</h1><p>Relatório de Projetos / Ordens de Serviço</p></div>
  <span class="badge">PROJETOS</span>
</div>

<div class="kpi-row">
  <div class="kpi slate"><label>Total de Projetos</label><span>{total}</span></div>
  <div class="kpi blue"><label>Em Andamento</label><span>{active}</span></div>
  <div class="kpi green"><label>Concluídos</label><span>{done}</span></div>
  <div class="kpi red"><label>Cancelados</label><span>{canc}</span></div>
</div>

<div class="kpi-row" style="grid-template-columns:repeat(3,1fr)">
  <div class="kpi green"><label>Receita Total</label><span>{_fmt(tot_rev)}</span></div>
  <div class="kpi amber"><label>Margem Média</label><span>{avg_margin:.1f}%</span></div>
  <div class="kpi blue"><label>Duração Média</label><span>{avg_duration:.0f} dias</span></div>
</div>

<div class="sec"><div class="sec-title">Lista de Projetos / OS ({total})</div>
<table><thead><tr>
  <th>Projeto / OS</th><th>Cliente</th><th>Status</th>
  <th style="text-align:right">Receita</th><th style="text-align:right">Custo</th>
  <th style="text-align:right">Margem</th><th>Início</th><th>Término</th>
</tr></thead>
<tbody>{proj_rows if proj_rows else '<tr><td colspan="8" style="text-align:center;padding:12px;color:#94A3B8">Nenhum projeto encontrado</td></tr>'}</tbody>
</table></div>

<div class="ftr"><span>InfraReport — Gestão de Infraestrutura Predial</span>
<span>Gerado em {date_cls.today().strftime("%d/%m/%Y")}</span></div>
</body></html>"""

    content = await _gotenberg(gotenberg_url, html)
    fname   = f"projetos_{date_cls.today().strftime('%Y%m%d')}.pdf"
    return _save(storage, fname, content)


# ── PROPOSTAS ─────────────────────────────────────────────────────────────────

async def pdf_propostas(
    gotenberg_url: str, storage: Path,
    user_id: str, proposals: list[dict],
    date_from: Optional[date_cls], date_to: Optional[date_cls],
) -> str:
    STATUS_PT = {"pendente":"Pendente","enviada":"Enviada","aprovada":"Aprovada","rejeitada":"Rejeitada"}
    STATUS_COL = {"pendente":"#F59E0B","enviada":"#3B82F6","aprovada":"#10B981","rejeitada":"#EF4444"}

    total    = len(proposals)
    approved = sum(1 for p in proposals if p.get("status") == "aprovada")
    rejected = sum(1 for p in proposals if p.get("status") == "rejeitada")
    sent     = sum(1 for p in proposals if p.get("status") == "enviada")
    conv_pct = approved / total * 100 if total > 0 else 0

    vals_aprov = [float(p.get("value") or 0) for p in proposals if p.get("status") == "aprovada" and p.get("value")]
    tot_aprov  = sum(vals_aprov)
    avg_aprov  = tot_aprov / len(vals_aprov) if vals_aprov else 0

    prop_rows = ""
    for i, p in enumerate(proposals):
        sc  = STATUS_COL.get(p.get("status",""), "#64748B")
        st  = STATUS_PT.get(p.get("status",""), "—")
        val = float(p.get("value") or 0)
        note = (p.get("notes") or p.get("rejection_reason") or "—")[:60]
        prop_rows += f"""<tr style="background:{_row(i)}">
          <td style="padding:5px 10px">{_d(p.get('created_at'))}</td>
          <td style="padding:5px 10px;font-weight:600">{(p.get('client_name') or '—')[:30]}</td>
          <td class="num" style="padding:5px 10px;color:#059669">{_fmt(val)}</td>
          <td style="padding:5px 10px"><span style="background:{sc}22;color:{sc};padding:2px 8px;border-radius:99px;font-size:10px;font-weight:700">{st}</span></td>
          <td style="padding:5px 10px;color:#64748B;font-size:10px">{note}</td>
        </tr>"""

    funnel_html = ""
    for status, count, color in [
        ("Criadas", total, "#64748B"),
        ("Enviadas", sent + approved + rejected, "#3B82F6"),
        ("Aprovadas", approved, "#10B981"),
    ]:
        pct = count / total * 100 if total > 0 else 0
        funnel_html += f"""<div style="margin-bottom:10px">
          <div style="display:flex;justify-content:space-between;margin-bottom:3px">
            <span style="font-size:11px;font-weight:600;color:#374151">{status}</span>
            <span style="font-size:11px;color:{color};font-weight:700">{count} ({pct:.0f}%)</span>
          </div>
          <div style="background:#F1F5F9;border-radius:4px;height:10px;overflow:hidden">
            <div style="width:{pct:.0f}%;height:100%;background:{color};border-radius:4px"></div>
          </div>
        </div>"""

    html = f"""<!DOCTYPE html><html lang="pt-BR"><head><meta charset="UTF-8"/>
<style>
{_base_css()}
.kpi-row{{grid-template-columns:repeat(4,1fr)}}
</style></head><body>
<div class="hdr">
  <div class="logo"><h1>InfraReport</h1><p>Relatório de Propostas Comerciais</p></div>
  <span class="badge">PROPOSTAS</span>
</div>

<div class="kpi-row">
  <div class="kpi slate"><label>Total de Propostas</label><span>{total}</span></div>
  <div class="kpi green"><label>Aprovadas</label><span>{approved}</span></div>
  <div class="kpi red"><label>Rejeitadas</label><span>{rejected}</span></div>
  <div class="kpi amber"><label>Taxa de Conversão</label><span>{conv_pct:.1f}%</span></div>
</div>

<div class="kpi-row" style="grid-template-columns:repeat(2,1fr)">
  <div class="kpi green"><label>Valor Total Aprovado</label><span>{_fmt(tot_aprov)}</span></div>
  <div class="kpi blue"><label>Ticket Médio (aprovadas)</label><span>{_fmt(avg_aprov)}</span></div>
</div>

<div class="sec"><div class="sec-title">Funil de Conversão</div>
{funnel_html}
</div>

<div class="sec"><div class="sec-title">Lista de Propostas ({total})</div>
<table><thead><tr>
  <th>Data</th><th>Cliente</th><th style="text-align:right">Valor</th>
  <th>Status</th><th>Observações</th>
</tr></thead>
<tbody>{prop_rows if prop_rows else '<tr><td colspan="5" style="text-align:center;padding:12px;color:#94A3B8">Nenhuma proposta encontrada</td></tr>'}</tbody>
</table></div>

<div class="ftr"><span>InfraReport — Gestão de Infraestrutura Predial</span>
<span>Gerado em {date_cls.today().strftime("%d/%m/%Y")}</span></div>
</body></html>"""

    content = await _gotenberg(gotenberg_url, html)
    fname   = f"propostas_{date_cls.today().strftime('%Y%m%d')}.pdf"
    return _save(storage, fname, content)


# ── EXCEL ─────────────────────────────────────────────────────────────────────

def _xl_header(ws, headers: list[str], fill_color: str = "1D4ED8") -> None:
    from openpyxl.styles import Font, PatternFill, Alignment
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _xl_autowidth(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)


def xlsx_financeiro(
    storage: Path, user_id: str, entries: list[dict],
    cat_map: dict[str, str],
    date_from: Optional[date_cls], date_to: Optional[date_cls],
) -> str:
    import openpyxl
    from openpyxl.styles import Font, Alignment, numbers

    wb = openpyxl.Workbook()

    # Sheet 1: Resumo Mensal
    ws1 = wb.active
    ws1.title = "Resumo Mensal"
    _xl_header(ws1, ["Mês", "Entradas (R$)", "Saídas (R$)", "Saldo (R$)", "Acumulado (R$)"])
    monthly: dict[str, dict] = {}
    for e in entries:
        key = str(e.get("date",""))[:7]
        if not key:
            continue
        r = monthly.setdefault(key, {"ent": 0.0, "sai": 0.0})
        if e.get("type") == "entrada":
            r["ent"] += float(e["value"])
        else:
            r["sai"] += float(e["value"])
    acc = 0.0
    for row_idx, key in enumerate(sorted(monthly), 2):
        y, m = key.split("-")
        ent = monthly[key]["ent"]; sai = monthly[key]["sai"]; sal = ent - sai
        acc += sal
        ws1.cell(row=row_idx, column=1, value=f"{MONTHS_PT[int(m)-1]}/{y}")
        ws1.cell(row=row_idx, column=2, value=round(ent, 2))
        ws1.cell(row=row_idx, column=3, value=round(sai, 2))
        ws1.cell(row=row_idx, column=4, value=round(sal, 2))
        ws1.cell(row=row_idx, column=5, value=round(acc, 2))
    _xl_autowidth(ws1)

    # Sheet 2: Lançamentos
    ws2 = wb.create_sheet("Lançamentos")
    _xl_header(ws2, ["Data","Tipo","Categoria","Descrição","Fornecedor","Forma Pag.","Valor (R$)"])
    for row_idx, e in enumerate(entries, 2):
        cat = cat_map.get(e.get("category",""), e.get("category",""))
        ws2.cell(row=row_idx, column=1, value=str(e.get("date",""))[:10])
        ws2.cell(row=row_idx, column=2, value="Entrada" if e.get("type")=="entrada" else "Saída")
        ws2.cell(row=row_idx, column=3, value=cat)
        ws2.cell(row=row_idx, column=4, value=e.get("description",""))
        ws2.cell(row=row_idx, column=5, value=e.get("supplier",""))
        ws2.cell(row=row_idx, column=6, value=e.get("payment_method",""))
        ws2.cell(row=row_idx, column=7, value=round(float(e.get("value",0)), 2))
    _xl_autowidth(ws2)

    df  = date_from or date_cls.today().replace(day=1)
    dt  = date_to   or date_cls.today()
    fname = f"financeiro_{df.strftime('%Y%m')}_{dt.strftime('%Y%m')}.xlsx"
    path  = storage / fname
    storage.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return str(path)


def xlsx_projetos(
    storage: Path, user_id: str, projects: list[dict],
    date_from: Optional[date_cls], date_to: Optional[date_cls],
) -> str:
    import openpyxl

    STATUS_PT = {"aguardando":"Aguardando","em_andamento":"Em Andamento","concluido":"Concluído","cancelado":"Cancelado"}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Projetos"
    _xl_header(ws, ["Nome","Cliente","Segmento","Status","Receita (R$)","Custo (R$)","Margem (%)","Início","Término","Duração (dias)"])

    for row_idx, p in enumerate(projects, 2):
        rev  = float(p.get("revenue") or 0)
        cost = float(p.get("material_cost",0) or 0) + float(p.get("labor_cost",0) or 0)
        mar  = round((rev - cost) / rev * 100, 1) if rev > 0 else 0
        dur  = ""
        if p.get("start_date") and p.get("end_date"):
            try:
                sd = date_cls.fromisoformat(str(p["start_date"])[:10])
                ed = date_cls.fromisoformat(str(p["end_date"])[:10])
                dur = (ed - sd).days
            except Exception:
                pass
        ws.cell(row=row_idx, column=1, value=p.get("name",""))
        ws.cell(row=row_idx, column=2, value=p.get("client",""))
        ws.cell(row=row_idx, column=3, value=p.get("segment",""))
        ws.cell(row=row_idx, column=4, value=STATUS_PT.get(p.get("status",""),""))
        ws.cell(row=row_idx, column=5, value=round(rev, 2))
        ws.cell(row=row_idx, column=6, value=round(cost, 2))
        ws.cell(row=row_idx, column=7, value=mar)
        ws.cell(row=row_idx, column=8, value=str(p.get("start_date",""))[:10])
        ws.cell(row=row_idx, column=9, value=str(p.get("end_date",""))[:10])
        ws.cell(row=row_idx, column=10, value=dur)
    _xl_autowidth(ws)

    fname = f"projetos_{date_cls.today().strftime('%Y%m%d')}.xlsx"
    path  = storage / fname
    storage.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return str(path)


def xlsx_propostas(
    storage: Path, user_id: str, proposals: list[dict],
    date_from: Optional[date_cls], date_to: Optional[date_cls],
) -> str:
    import openpyxl

    STATUS_PT = {"pendente":"Pendente","enviada":"Enviada","aprovada":"Aprovada","rejeitada":"Rejeitada"}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Propostas"
    _xl_header(ws, ["Data","Cliente","Valor (R$)","Status","Observações"])

    for row_idx, p in enumerate(proposals, 2):
        ws.cell(row=row_idx, column=1, value=str(p.get("created_at",""))[:10])
        ws.cell(row=row_idx, column=2, value=p.get("client_name",""))
        ws.cell(row=row_idx, column=3, value=round(float(p.get("value") or 0), 2))
        ws.cell(row=row_idx, column=4, value=STATUS_PT.get(p.get("status",""),""))
        ws.cell(row=row_idx, column=5, value=p.get("notes") or p.get("rejection_reason") or "")
    _xl_autowidth(ws)

    # Sheet 2: Análise
    ws2 = wb.create_sheet("Análise")
    _xl_header(ws2, ["Métrica","Valor"], fill_color="059669")
    total    = len(proposals)
    approved = sum(1 for p in proposals if p.get("status") == "aprovada")
    rejected = sum(1 for p in proposals if p.get("status") == "rejeitada")
    sent     = sum(1 for p in proposals if p.get("status") == "enviada")
    vals_a   = [float(p.get("value") or 0) for p in proposals if p.get("status") == "aprovada" and p.get("value")]
    tot_a    = sum(vals_a)
    metrics = [
        ("Total de Propostas",      total),
        ("Aprovadas",               approved),
        ("Rejeitadas",              rejected),
        ("Enviadas (pendente resp)", sent),
        ("Taxa de Conversão (%)",   round(approved / total * 100, 1) if total else 0),
        ("Valor Total Aprovado (R$)", round(tot_a, 2)),
        ("Ticket Médio Aprovado (R$)", round(tot_a / len(vals_a), 2) if vals_a else 0),
    ]
    for row_idx, (label, value) in enumerate(metrics, 2):
        ws2.cell(row=row_idx, column=1, value=label)
        ws2.cell(row=row_idx, column=2, value=value)
    _xl_autowidth(ws2)

    fname = f"propostas_{date_cls.today().strftime('%Y%m%d')}.xlsx"
    path  = storage / fname
    storage.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return str(path)
